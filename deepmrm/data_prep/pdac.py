from pathlib import Path
import xml.etree.ElementTree as ET

import joblib
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pyopenms import AASequence, Residue, MzMLFile
from deepmrm.data.mrm_experiment import MRMExperiment
from deepmrm import data_dir, get_yaml_config, private_data_dir
from deepmrm.data.dataset import MRMDataset
from deepmrm.data.transition import TransitionData
from mstorch.data.mass_spec.tolerance import Tolerance
from mstorch.utils.logger import get_logger

logger = get_logger('DeepMRM')
_conf = get_yaml_config()
conf = _conf['PDAC-SIT']

PDAC_SIT_DIR = Path(conf['ROOT_DIR'])
PDAC_SIT_MS_DIR = Path(conf['MZML_DIR'])
PDAC_SIT_MANUAL_DIR = Path(conf['MANUAL_DIR'])
SKYLINE_DIR = Path(conf['SKYLINE_DIR'])
TRANSITION_FILE = conf['TRANSITION_FILE']

def get_msdata_df():
    """Scan file list for PDAC-SIT dataset, and extract metadata from filenames
    Returns:
        pd.DataFrame: file list with metadata
    """
    save_path = Path(data_dir / 'PDAC_SIT_files.csv')
    
    if not save_path.exists():
        ms_files = list()
        for fname in PDAC_SIT_MS_DIR.iterdir():
            if fname.is_file():
                file_name = fname.name
                s = file_name.split('_')
                patient_id = int(s[1])
                replicate_id = int(s[2])
                ms_files.append([patient_id, replicate_id, file_name])

        df = pd.DataFrame(ms_files, columns=['patient_id', 'replicate_id', 'mzML'])
        df = df.sort_values(['patient_id', 'replicate_id']).reset_index(drop=True)

        # reset replicate id (one-based index)
        for grp, sub_df in df.groupby('patient_id'):
            df.loc[sub_df.index, 'replicate_id'] = np.arange(1, sub_df.shape[0]+1)

        df.to_csv(save_path, index=False)
    else:
        df = pd.read_csv(save_path)

    return df

def _get_raw_transition_df():
    tran_df = pd.read_excel(private_data_dir / TRANSITION_FILE)
    peptide_sequence = tran_df['Compound Name'].replace({
            '\[pS\]': 'S(Phospho)', 
            '\[pT\]': 'T(Phospho)',
            '\(IAA\)': '(Carbamidomethyl)',
            '.heavy': '(Label:13C(6))',
            '.light': ''
            }, regex=True).apply(AASequence.fromString)

    precursor_mass = peptide_sequence.apply(lambda seq: seq.getMonoWeight(Residue.ResidueType.Full, 2))
    tran_df['Sequence'] = peptide_sequence
    tran_df['Precursor Charge'] = (precursor_mass / tran_df['Precursor Ion']).round().astype(int)
    return tran_df


def get_transition_df():
    """Read a transition list in excel file, and convert it to DataFrame
    Returns:
        [type]: [description]
    """
    save_path = Path(data_dir / 'PDAC_SIT_transition.csv')

    if not save_path.exists():
        tran_df = _get_raw_transition_df()
        cols_rename = {
            'Compound Group': 'peptide_id', 
            'ISTD?': 'is_heavy', 
            'Precursor Ion': 'precursor_mz', 
            'Product Ion': 'product_mz', 
            'Ret Time (min)': 'rt',
            'Precursor Charge': 'precursor_charge'
        }

        tran_df = tran_df[list(cols_rename)].rename(columns=cols_rename)
        tran_df.to_csv(save_path, index=False)
    else:
        tran_df = pd.read_csv(save_path)

    return tran_df


# def get_compact_transition_df():
    
#     trans_df = get_transition_df()
#     compact_trans_df = trans_df.drop_duplicates(
#                         ['Compound Group', 'Compound Name', 'Ret Time (min)'], 
#                         keep='first')
#     selected_cols = [
#         'Compound Group', 'Compound Name', 'Ret Time (min)',
#         'Sequence', 'Heavy'
#     ]

#     compact_trans_df =compact_trans_df[selected_cols].reset_index(drop=True)
#     compact_trans_df.index.name = 'compound_idx'

#     return compact_trans_df



def _extract_manual_data(df, sheet):
    df = df.replace({'X': np.nan, '-': np.nan})
    headers = [header_col[0] for header_col in df.columns]
    ratio_index = headers.index('Final Endo/SIL ratio after inspection')
    
    all_records = []
    for i in range(0, df.shape[0], 3):
        peptide_code = df.iloc[i, 0]
        if pd.isna(peptide_code):
            continue

        cs2_cell_color = sheet[f'D{4+i}'].font.color
        cs3_cell_color = sheet[f'E{4+i}'].font.color
        if (cs2_cell_color and cs2_cell_color.index == 'FFFF0000'):
            selected_charge = 2
        elif (cs3_cell_color and cs3_cell_color.index == 'FFFF0000'):
            selected_charge = 3
        else:
            selected_charge = 0

        trans_quality = [
            sheet[f'R{4+i+k}'].fill.bgColor.value == 64 
                for k in range(2, -1, -1)
        ]

        light_frag_auc = df.iloc[i:i+3, 6].values[::-1]
        # light_frag_rank = df.iloc[i:i+3, 7].values[::-1]
        light_frag_rt = df.iloc[i:i+3, 8].values[::-1]
        light_rt = np.nanmean(light_frag_rt)

        heavy_frag_auc = df.iloc[i:i+3, 12].values[::-1]
        # heavy_frag_rank = df.iloc[i:i+3, 13].values[::-1]
        heavy_frag_rt = df.iloc[i:i+3, 14].values[::-1]
        heavy_rt = np.nanmean(heavy_frag_rt)

        manual_frag_ratio = light_frag_auc / heavy_frag_auc

        light_auc = df.iloc[i, 9]
        heavy_auc = df.iloc[i, 15]
        ion_order = df.iloc[i, 20]
        if not isinstance(ion_order, bool):
            ion_order = ion_order.lower().find('true') > 0

        manual_ratio = df.iloc[i, ratio_index]
        manual_ratio_desc = df.iloc[i, ratio_index+1]
        heavy_pmol = df.iloc[i, ratio_index+2]
        light_pmol = df.iloc[i, ratio_index+3]
        
        record = [
            peptide_code, selected_charge, 
            light_auc, heavy_auc, ion_order, manual_ratio, 
            manual_ratio_desc, light_pmol, heavy_pmol,
            light_rt, heavy_rt,
        ]
        record.extend(list(manual_frag_ratio))
        record.extend(trans_quality)
        record.extend(list(light_frag_auc))
        record.extend(list(heavy_frag_auc))

        all_records.append(record)

    cols = [
        'peptide_code', 'selected_charge', 
        'light_auc', 'heavy_auc', 'ion_order', 
        'manual_ratio', 'manual_ratio_desc', 
        'light_pmol', 'heavy_pmol', 
        'light_rt', 'heavy_rt'
    ]
    cols.extend([f'manual_frag_ratio_t{k+1}' for k in range(3)])
    cols.extend([f'manual_frag_quality_t{k+1}' for k in range(3)])
    cols.extend([f'manual_light_frag_auc_t{k+1}' for k in range(3)])
    cols.extend([f'manual_heavy_frag_auc_t{k+1}' for k in range(3)])

    return pd.DataFrame(all_records, columns=cols)


def _create_label_df():
    dfs = []
    for fname in PDAC_SIT_MANUAL_DIR.iterdir():
        if not fname.is_file() or not fname.name.startswith('PDAC'):
            continue
        
        # fname = PDAC_SIT_MANUAL_DIR / 'PDAC273_Sub6_spiked_ratio_rep1-3_outliertest_20211020.xlsx'
        print(fname.name)
        patient_id = int(fname.name[4:7])
        
        wb = load_workbook(fname)
        sheet_names = wb.sheetnames

        for sheet_idx in range(3):
            sh_name = sheet_names[sheet_idx]
            df = pd.read_excel(fname, sheet_name=sheet_idx, header=[0, 1, 2])
            sh = wb[sh_name]
            df = _extract_manual_data(df, sh)
            df['patient_id'] = patient_id
            df['replicate_id'] = sheet_idx + 1
            dfs.append(df)

    label_df = pd.concat(dfs, ignore_index=True)

    return label_df

def get_label_df():
    fpath = data_dir / 'PDAC_SIT_label.csv'
    if fpath.exists():
        label_df = pd.read_csv(fpath)
    else:
        label_df = _create_label_df()
        # manual correction (Excel file has been updated)
        # m = (label_df['patient_id'] == 211) & (label_df['replicate_id'] == 1) & \
        #     (label_df['peptide_code'] == 'PDAC0097')
        # idx = label_df.index[m][0]
        # label_df.loc[idx, 'manual_ratio'] = (30746+2254)/(313490+43943)

        # m = (label_df['patient_id'] == 186) & (label_df['replicate_id'] == 1) & \
        #     (label_df['peptide_code'] == 'PDAC0168')
        # idx = label_df.index[m][0]        
        # label_df.loc[idx, 'manual_ratio_desc'] = 'y ion transition 순서 불일치'
        label_df.to_csv(fpath, index=False)

    ms_df = get_msdata_df()
    label_df = label_df.merge(ms_df, 
                            on=['patient_id', 'replicate_id'], 
                            how='inner')\
                       .reset_index(drop=True)\
                       .rename(columns={'peptide_code': 'peptide_id'})
    label_df.index.name = 'label_idx'

    return label_df


def _create_chrom_df():
    # save_path = private_data_dir / 'PDAC_SIT_chrom.pkl'
    # chrom_df = pd.read_pickle(save_path)
    save_path = data_dir / 'PDAC_SIT_xic.pkl'
    ms_df = get_msdata_df()
    trans_df = get_transition_df()
    label_df = get_label_df()

    xic_data = dict()
    for file_id, row in ms_df.iterrows():
        print(f'File index: {file_id}')
        mzml_file = row['mzML']
        patient_id = row['patient_id']
        replicate_id = row['replicate_id']

        mzml_path = PDAC_SIT_MS_DIR / mzml_file
        
        label_mask = (label_df['patient_id'] == patient_id) & (label_df['replicate_id'] == replicate_id)
        temp_label_df = label_df[label_mask]
        # pep_id_to_label_idx_map = temp_label_df[['peptide_code']].reset_index(drop=False).set_index('peptide_code')

        if np.sum(label_mask) == 0:
            # skip no label data
            logger.info(f'{mzml_file} is not annotated..Skip')
            continue

        temp_trans_df = trans_df.merge(
                            temp_label_df[['peptide_id', 'selected_charge', 'light_rt']],
                            left_on=['peptide_id', 'precursor_charge'],
                            right_on=['peptide_id', 'selected_charge'],
                            how='inner'
                        )

        trans_data = TransitionData(temp_trans_df, rt_col='light_rt')
        ds = MRMDataset(
            file_path=mzml_path,
            transition_data=trans_data,
        )
        tolerance = Tolerance(100)
        ds.extract_data(tolerance, filter_by_rt=True)

        for i in range(len(ds)):
            sample = ds[i]
            pep_id = sample['peptide_id']
            # label_idx = pep_id_to_label_idx_map.at[pep_id, 'label_idx']
            key = (patient_id, replicate_id, pep_id)
            xic_data[key] = sample['XIC']

    joblib.dump(xic_data, save_path)


def get_xic_data():
    save_path = data_dir / 'PDAC_SIT_xic.pkl'
    return joblib.load(save_path)


def get_skyline_df():

    save_path = data_dir / 'PDAC_SIT_skyline.csv'

    if not save_path.exists():

        sky_df = _get_skyline_df()
        trans_df = pd.read_excel(private_data_dir / TRANSITION_FILE)
        peptide_sequence = trans_df['Compound Name'].replace({
                    '\[pS\]': 'S(Phospho)', 
                    '\[pT\]': 'T(Phospho)',
                    '\(IAA\)': '(Carbamidomethyl)',
                    '.heavy': '(Label:13C(6))',
                    '.light': ''
                    }, regex=True).apply(AASequence.fromString)
        trans_df['plain_seq'] = peptide_sequence.apply(lambda x : x.toUnmodifiedString())

        assert trans_df['plain_seq'].nunique() == trans_df['Compound Group'].nunique()
        
        seq_to_pep_id = trans_df.drop_duplicates(['plain_seq', 'Compound Group'], keep='first')[['plain_seq', 'Compound Group']]
        seq_to_pep_id = seq_to_pep_id.rename(columns={'Compound Group': 'peptide_id'})
        sky_df = sky_df.merge(seq_to_pep_id, left_on='sequence', right_on='plain_seq', how='inner').drop(columns=['plain_seq', 'sequence'])
        sky_df.to_csv(save_path, index=False)

    else:
        sky_df = pd.read_csv(save_path)

    return sky_df



def _get_skyline_df():

    save_path = private_data_dir / 'PDAC_SIT_skyline.csv'
    def extract_pid_rid(x):
        s = x.split('_')
        pid = s[1]
        rid = s[2]
        if rid == 'Nano':
            rid = s[-2]
        return int(pid), int(rid)

    if save_path.exists():
        df = pd.read_csv(save_path)
    else:
        results = list()
        for sky_path in SKYLINE_DIR.rglob('*.sky'):
            # fname = 'PDAC_Verification_PDAC73_2_Sub4_060821.sky'
            # fname = 'PDAC_Verification_254_Re_Sub6.sky'
            # sky_path = SKYLINE_DIR/fname
            tree = ET.parse(sky_path)
            root = tree.getroot()
            rep_name = root.find('settings_summary').find('measured_results').find('replicate').get('name')

            if sky_path.stem == 'PDAC_153_Rep2_Target_031121_updated':
                pid, rid = 193, 2
            elif sky_path.stem == 'PDAC_153_Rep3_Target_031121_updated':
                pid, rid = 193, 3
            elif sky_path.stem == 'PDAC_Verification_254_Re_Sub6':
                pid, rid = 254, 2
            else:
                pid, rid = extract_pid_rid(rep_name)
            
            for peptide in root.iter('peptide'):
                sequence = peptide.get('sequence')
                precursor = peptide.find('precursor')
                precursor_peak = precursor.find('precursor_results').find('precursor_peak')
                info = {'patient_id': pid, 'replicate_id': rid, 'sequence': sequence}
                info.update(precursor_peak.attrib)
                results.append(info)
        df = pd.DataFrame.from_dict(results)
        df.to_csv(save_path, index=False) 

    float_cols = [
        'peak_count_ratio', 'retention_time', 'start_time',
        'end_time', 'fwhm', 'area', 'background', 'height'
    ]
    for col in float_cols:
        df[col] = df[col].astype(np.float64)
    
    df['patient_id'] = df['patient_id'].astype(int)
    df['replicate_id'] = df['replicate_id'].astype(int)
    df = df.sort_values(['patient_id', 'replicate_id', 'sequence'])

    # reset replicate id (one-based index)
    for grp, sub_df in df.groupby('patient_id'):
        d = {rid: i+1 for i, rid in enumerate(sub_df['replicate_id'].unique())}
        if len(d) != 3:
            ValueError('something wrong')
        df.loc[sub_df.index, 'replicate_id'] = sub_df['replicate_id'].replace(d)

    return df


def get_metadata_df():

    xic_data = get_xic_data()
    label_df = get_label_df()
    sky_df = get_skyline_df()

    cols = ['patient_id', 'replicate_id', 'peptide_id', 'start_time', 'end_time']
    label_df = label_df.reset_index()\
                       .merge(
                            sky_df[cols], 
                            how='inner',
                            on=cols[:3])\
                       .set_index('label_idx')
    
    label_df['manual_quality'] = label_df['manual_ratio'].notnull().astype(np.int64)

    return label_df, xic_data

